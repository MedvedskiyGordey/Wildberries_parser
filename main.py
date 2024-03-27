import asyncio
import json
import logging

import coloredlogs as coloredlogs
import openpyxl

import requests as requests
import telebot

coloredlogs.install(level="DEBUG")

bot = telebot.TeleBot('XXXXX:YYYYYY')  # токен бота
chat_id = -0000000000000  # id группы

count_messages = 1  # кол-во последних отзывов товара

delay = 3600  # промежуток в секундах


@bot.message_handler(content_types=['text'])
def start(message):
    bot.send_message(message.chat.id, message.text)


async def start_bot():
    log = logging.getLogger('start_bot')
    try:
        bot.send_message(chat_id, 'bot start')

        while True:
            feedback = check_feedback()
            for item in feedback:
                for mes in item['messages'][-count_messages:]:
                    bot.send_message(chat_id, mes['message'])
            await asyncio.sleep(delay)
    except asyncio.CancelledError:
        logging.warning('Принудительное завершение')
    except Exception as error:
        log.error('Ошибка > %s', error)


def get_wb_imtid(wb_id):
    log = logging.getLogger(f'get_wb_imtid')
    product = {
        "imtid": None,
        "id": wb_id,
        "name": None
    }
    try:
        target_url = f"https://card.wb.ru/cards/v2/detail?appType=1&curr=rub&dest=-1257786&spp=30&nm={wb_id}"

        resp = requests.get(target_url)
        search_json = json.loads(resp.text)

        products_list = search_json.get('data', {}).get('products', [])

        if products_list:
            product['imtid'] = products_list[0].get('root', None)
            product['name'] = products_list[0].get('name', None)
    except Exception as error:
        log.error('Ошибка > %s', error)
    finally:
        return product


def get_feedback(product):
    log = logging.getLogger(f'get_feedback')
    messages_dict = {
        'product_id': product.get('id', None),
        'messages': []
    }
    try:
        target_url = f"https://feedbacks1.wb.ru/feedbacks/v1/{product['imtid']}"
        resp = requests.get(target_url)
        search_json = json.loads(resp.text)
        feedbacks_list = search_json.get('feedbacks', [])
        valuation = search_json.get('valuation', None)
        if feedbacks_list:
            sorted_feedbacks_list = sorted(feedbacks_list, key=lambda d: d['createdDate'], reverse=True)

            for item in sorted_feedbacks_list:
                productValuation = item.get('productValuation', None)
                if productValuation is not None and 1 <= productValuation <= 4:
                    messages_dict['messages'].append({
                        'message': f"Негативный отзыв/{product['name']}/{item.get('nmId', None)}/{item.get('productValuation', None)}/ {item.get('text', None)}/{valuation}.",
                        'date': item.get('createdDate')
                    })
    except Exception as error:
        log.error('Ошибка > %s', error)
    finally:
        return messages_dict


def read_excel(path):
    products = []
    wookbook = openpyxl.load_workbook(path)
    worksheet = wookbook.active
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            products.append(col[i].value)
    return products


def check_feedback():
    products = read_excel("product_id.xlsx")
    feedback = []

    for item in products:
        product = get_wb_imtid(item)
        feedback_list = get_feedback(product)
        feedback.append(feedback_list)
    return feedback


if __name__ == '__main__':
    asyncio.run(start_bot())
    bot.polling(none_stop=True, interval=0)
